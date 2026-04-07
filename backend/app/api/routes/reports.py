"""
REST endpoints for report generation.

GET /reports/scenario/{id}/excel  – Download a scenario report as .xlsx
GET /reports/scenario/{id}/json   – Download full scenario data as JSON

Both endpoints accept an optional ?db= query parameter to select the target
database (e.g. ?db=lip2_belgium).  This is needed because the download links
are plain <a href> tags that cannot send the X-LIP2-Database header.
"""

import io
import json
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import get_settings
from app.database import get_session_factory
from app.dependencies import get_db, get_db_name
from app.models.census import CensusArea
from app.models.optimization import OptimizationResult, OptimizationScenario

router = APIRouter(prefix="/reports", tags=["reports"])


async def _resolve_db(
    db_query: str | None,
    db_from_header: AsyncSession,
) -> AsyncSession:
    """
    Return the correct async session.

    Priority: ?db= query param > X-LIP2-Database header > default database.
    When the query param is present a new session is opened directly; the
    caller is responsible for closing it.
    """
    if db_query is None:
        return db_from_header, False   # (session, owned_by_us)

    available = get_settings().available_databases
    if db_query not in available:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown database '{db_query}'. Available: {available}",
        )
    factory = get_session_factory(db_query)
    session = factory()
    return session, True   # (session, owned_by_us)


@router.get("/scenario/{scenario_id}/excel")
async def export_scenario_excel(
    scenario_id: int,
    db_query: str | None = Query(default=None, alias="db"),
    db: AsyncSession = Depends(get_db),
):
    """Export scenario results as an Excel workbook (.xlsx)."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl is not installed")

    session, owned = await _resolve_db(db_query, db)
    try:
        scenario = await session.get(OptimizationScenario, scenario_id)
        if not scenario:
            raise HTTPException(status_code=404, detail="Scenario not found")

        res = await session.execute(
            select(OptimizationResult, CensusArea)
            .join(CensusArea, OptimizationResult.census_area_id == CensusArea.id)
            .where(OptimizationResult.scenario_id == scenario_id)
        )
        rows = res.all()
        
        # Pre-load census areas for the served-areas sheet.
        all_area_ids: set[int] = set()
        for opt_res, _ in rows:
            for entry in (opt_res.served_area_ids or []):
                all_area_ids.add(int(entry[0]) if isinstance(entry, (list, tuple)) else int(entry))
        area_map: dict[int, CensusArea] = {}
        if all_area_ids:
            area_res = await session.execute(
                select(CensusArea).where(CensusArea.id.in_(list(all_area_ids)))
            )
            area_map = {a.id: a for a in area_res.scalars().all()}
        
        # Pre-load census areas for unassigned areas sheet.
        unassigned_raw = (scenario.result_stats or {}).get("_unassigned_areas", [])
        uncov_area_ids = [ua["census_area_id"] for ua in unassigned_raw if ua.get("census_area_id")]
        uncov_area_map: dict[int, CensusArea] = {}
        if uncov_area_ids:
            uncov_res = await session.execute(
                select(CensusArea).where(CensusArea.id.in_(uncov_area_ids))
            )
            uncov_area_map = {a.id: a for a in uncov_res.scalars().all()}
    finally:
        if owned:
            await session.close()

    wb = openpyxl.Workbook()

    # --- Sheet 1: Scenario Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")

    ws_summary.append(["LIP2 – Optimization Scenario Report"])
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.append([])
    ws_summary.append(["Scenario ID", scenario.id])
    ws_summary.append(["Name", scenario.name])
    ws_summary.append(["Model", scenario.model_type.value])
    ws_summary.append(["Facilities (p)", scenario.p_facilities])
    ws_summary.append(["Service Radius (min)", scenario.service_radius or "N/A"])
    ws_summary.append(["Status", scenario.status.value])
    ws_summary.append(["Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")])
    ws_summary.append([])

    if scenario.result_stats:
        ws_summary.append(["── Statistics ──"])
        for key, value in scenario.result_stats.items():
            if key.startswith("_"):
                continue
            if not isinstance(value, (int, float, str, bool, type(None))):
                value = str(value)
            ws_summary.append([key.replace("_", " ").title(), value])

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 25

    # --- Sheet 2: Facility Locations ---
    ws_fac = wb.create_sheet("Facility Locations")
    fac_headers = ["#", "Area Code", "Area Name", "Province", "Canton", "Parish",
                   "X (Lon)", "Y (Lat)", "Demand Covered"]
    ws_fac.append(fac_headers)
    for cell in ws_fac[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, (opt_res, area) in enumerate(rows, start=1):
        ws_fac.append([
            i,
            area.area_code,
            area.name or "",
            area.province_code,
            area.canton_code,
            area.parish_code,
            area.x,
            area.y,
            round(opt_res.covered_demand or 0.0, 2),
        ])

    for col in ws_fac.columns:
        max_len = max(len(str(cell.value or "")) for cell in col) + 2
        ws_fac.column_dimensions[col[0].column_letter].width = min(max_len, 30)
    
    # --- Sheet 3: Served Areas ---
    ws_served = wb.create_sheet("Served Areas")
    served_headers = [
        "Facility #", "Facility Code", "Facility Name",
        "Area Code", "Area Name", "Province", "Canton", "Parish",
        "X (Lon)", "Y (Lat)", "Demand Assigned",
    ]
    ws_served.append(served_headers)
    for cell in ws_served[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    for fac_num, (opt_res, fac_area) in enumerate(rows, start=1):
        for entry in (opt_res.served_area_ids or []):
            if isinstance(entry, (list, tuple)):
                aid, demand_amt = int(entry[0]), float(entry[1])
            else:
                aid, demand_amt = int(entry), 0.0
            sa = area_map.get(aid)
            if sa is None:
                continue
            ws_served.append([
                fac_num,
                fac_area.area_code,
                fac_area.name or "",
                sa.area_code,
                sa.name or "",
                sa.province_code,
                sa.canton_code,
                sa.parish_code,
                sa.x,
                sa.y,
                round(demand_amt, 2),
            ])

    for col in ws_served.columns:
        max_len = max(len(str(cell.value or "")) for cell in col) + 2
        ws_served.column_dimensions[col[0].column_letter].width = min(max_len, 30)

    # --- Sheet 4: Uncovered Areas ---
    ws_uncov = wb.create_sheet("Uncovered Areas")
    uncov_headers = [
        "Area Code", "Area Name", "Province", "Canton", "Parish",
        "X (Lon)", "Y (Lat)", "Demand",
        "Nearest Facility", "Travel Time (min)", "Distance (km)",
    ]
    ws_uncov.append(uncov_headers)
    uncov_header_fill = PatternFill(fill_type="solid", fgColor="8B1A1A")
    for cell in ws_uncov[1]:
        cell.font = header_font
        cell.fill = uncov_header_fill
        cell.alignment = Alignment(horizontal="center")

    for ua in unassigned_raw:
        aid = ua.get("census_area_id")
        ca  = uncov_area_map.get(aid) if aid else None
        ws_uncov.append([
            ua.get("area_code") or (ca.area_code if ca else ""),
            ua.get("name") or (ca.name if ca else ""),
            ca.province_code if ca else "",
            ca.canton_code  if ca else "",
            ca.parish_code  if ca else "",
            ua.get("x") or (ca.x if ca else None),
            ua.get("y") or (ca.y if ca else None),
            round(ua.get("demand") or 0.0, 2),
            ua.get("nearest_facility_code") or "",
            round(ua.get("nearest_facility_travel_time_min") or 0.0, 2),
            round(ua.get("nearest_facility_distance_km") or 0.0, 2),
        ])

    for col in ws_uncov.columns:
        max_len = max(len(str(cell.value or "")) for cell in col) + 2
        ws_uncov.column_dimensions[col[0].column_letter].width = min(max_len, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"lip2_scenario_{scenario_id}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/scenario/{scenario_id}/json")
async def export_scenario_json(
    scenario_id: int,
    db_query: str | None = Query(default=None, alias="db"),
    db: AsyncSession = Depends(get_db),
):
    """Export full scenario data as a JSON download."""
    session, owned = await _resolve_db(db_query, db)
    try:
        scenario = await session.get(OptimizationScenario, scenario_id)
        if not scenario:
            raise HTTPException(status_code=404, detail="Scenario not found")

        res = await session.execute(
            select(OptimizationResult, CensusArea)
            .join(CensusArea, OptimizationResult.census_area_id == CensusArea.id)
            .where(OptimizationResult.scenario_id == scenario_id)
        )
        rows = res.all()
    finally:
        if owned:
            await session.close()

    payload = {
        "scenario": {
            "id": scenario.id,
            "name": scenario.name,
            "model_type": scenario.model_type.value,
            "p_facilities": scenario.p_facilities,
            "service_radius": scenario.service_radius,
            "status": scenario.status.value,
            "stats": scenario.result_stats,
            "created_at": scenario.created_at.isoformat() if scenario.created_at else None,
        },
        "facility_locations": [
            {
                "area_id": area.id,
                "area_code": area.area_code,
                "name": area.name,
                "province_code": area.province_code,
                "canton_code": area.canton_code,
                "parish_code": area.parish_code,
                "x": area.x,
                "y": area.y,
                "covered_demand": opt_res.covered_demand,
            }
            for opt_res, area in rows
        ],
    }

    content = json.dumps(payload, indent=2, ensure_ascii=False)
    filename = f"lip2_scenario_{scenario_id}.json"
    return StreamingResponse(
        iter([content]),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
